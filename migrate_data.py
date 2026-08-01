import sys, os, json, sqlite3

import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), 'registro.db')

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_conn()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, description TEXT);
        CREATE TABLE IF NOT EXISTS products (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, category_id INTEGER, brand TEXT, model TEXT, variant TEXT, compatibility TEXT, price_cost REAL DEFAULT 0, price_sale REAL DEFAULT 0, stock INTEGER DEFAULT 0, min_stock INTEGER DEFAULT 0, created_at TEXT DEFAULT (datetime('now','localtime')), updated_at TEXT DEFAULT (datetime('now','localtime')));
        CREATE TABLE IF NOT EXISTS sales (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT DEFAULT (datetime('now','localtime')), product_id INTEGER, product_name TEXT, quantity INTEGER DEFAULT 1, unit_price REAL, total REAL, payment_method TEXT, client_name TEXT, notes TEXT);
        CREATE TABLE IF NOT EXISTS services (id INTEGER PRIMARY KEY AUTOINCREMENT, order_num TEXT UNIQUE, date_in TEXT DEFAULT (datetime('now','localtime')), client TEXT, phone TEXT, model TEXT, fault TEXT, amount REAL DEFAULT 0, payment_method TEXT, date_out TEXT, status TEXT DEFAULT 'Por entregar', observations TEXT);
        CREATE TABLE IF NOT EXISTS inventory_movements (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT DEFAULT (datetime('now','localtime')), product_id INTEGER, type TEXT, quantity INTEGER, reason TEXT, reference TEXT);
        CREATE TABLE IF NOT EXISTS payment_methods (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE);
        CREATE TABLE IF NOT EXISTS service_statuses (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE);
    ''')
    defaults = [
        "INSERT OR IGNORE INTO categories (name) VALUES ('Pantalla')",
        "INSERT OR IGNORE INTO categories (name) VALUES ('Telefono')",
        "INSERT OR IGNORE INTO categories (name) VALUES ('Accesorio')",
        "INSERT OR IGNORE INTO categories (name) VALUES ('Repuesto')",
        "INSERT OR IGNORE INTO payment_methods (name) VALUES ('Divisas (USD Cash)')",
        "INSERT OR IGNORE INTO payment_methods (name) VALUES ('Pago Movil')",
        "INSERT OR IGNORE INTO payment_methods (name) VALUES ('Punto de Venta')",
        "INSERT OR IGNORE INTO payment_methods (name) VALUES ('Transferencia Zelle')",
        "INSERT OR IGNORE INTO payment_methods (name) VALUES ('Transferencia Bs')",
        "INSERT OR IGNORE INTO payment_methods (name) VALUES ('Efectivo Bs')",
        "INSERT OR IGNORE INTO service_statuses (name) VALUES ('Por entregar')",
        "INSERT OR IGNORE INTO service_statuses (name) VALUES ('Reparado / Pendiente Pago')",
        "INSERT OR IGNORE INTO service_statuses (name) VALUES ('Entregado')",
        "INSERT OR IGNORE INTO service_statuses (name) VALUES ('Cancelado / Devuelto')",
    ]
    for q in defaults:
        conn.execute(q)
    conn.commit()
    return conn

def add_service(conn, order, client, phone, model, fault, amount, payment, obs):
    try:
        conn.execute("INSERT INTO services (order_num, date_in, client, phone, model, fault, amount, payment_method, observations) VALUES (?, date('now'), ?,?,?,?,?,?,?)",
                     (order, client, phone, model, fault, amount, payment, obs))
        return True
    except:
        return False

def import_from_excel(conn, path, label):
    if not os.path.exists(path):
        print(f"  {label}: not found")
        return 0
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Registro_Reparaciones' not in wb.sheetnames:
        print(f"  {label}: no Registro_Reparaciones sheet")
        return 0
    ws = wb['Registro_Reparaciones']
    count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4 or not row[0] or not str(row[0]).startswith('ORD-'):
            continue
        order = str(row[0]).strip()
        client = str(row[2]).strip() if row[2] else ''
        phone = str(row[3]).strip() if row[3] else ''
        model = str(row[4]).strip() if row[4] else ''
        fault = str(row[5]).strip() if row[5] else ''
        amount = float(row[6]) if row[6] else 0
        payment = str(row[7]).strip() if row[7] else 'Divisas (USD Cash)'
        obs = str(row[10]).strip() if row[10] else ''
        if add_service(conn, order, client, phone, model, fault, amount, payment, obs):
            count += 1
    print(f"  {label}: {count} services imported")
    return count

def export_to_json():
    conn = get_conn()
    data = {}
    for table in ['categories', 'payment_methods', 'service_statuses', 'products', 'sales', 'services', 'inventory_movements']:
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()
        data[table] = [dict(r) for r in rows]
    conn.close()
    out_path = os.path.join(os.path.dirname(__file__), 'registro_export.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON export: {out_path}")

def import_products(conn):
    products = [
        ("Pantalla Red Note 11 / Note 11S 4G", 1, "Xiaomi", "Red Note 11", "", '["Red Note 11", "Note 11S 4G"]', 0, 0, 3, 2),
        ("Pantalla Red 10 4G", 1, "Xiaomi", "Red 10 4G", "", '[]', 0, 0, 2, 2),
        ("Pantalla Red Note 8", 1, "Xiaomi", "Red Note 8", "", '[]', 0, 0, 5, 2),
        ("Pantalla Red 10C", 1, "Xiaomi", "Red 10C", "", '[]', 0, 0, 5, 2),
        ("Pantalla Red A1 / A2", 1, "Xiaomi", "Red A1", "", '["Red A1", "Red A2"]', 0, 0, 3, 2),
        ("Pantalla Samsung A32 4G", 1, "Samsung", "A32 4G", "", '[]', 0, 0, 2, 2),
        ("Pantalla Samsung A06 4G", 1, "Samsung", "A06 4G", "", '[]', 0, 0, 5, 2),
        ("Pantalla Inf Hot 40i / Go 2024 / Spark 20", 1, "Infinix", "Hot 40i", "", '["Hot 40i", "Go 2024", "Spark 20"]', 0, 0, 9, 2),
        ("Pantalla Red 13C", 1, "Xiaomi", "Red 13C", "", '[]', 0, 0, 5, 2),
        ("Pantalla Red 14C", 1, "Xiaomi", "Red 14C", "", '[]', 0, 0, 3, 2),
        ("Pantalla Red A3", 1, "Xiaomi", "Red A3", "", '[]', 0, 0, 10, 2),
        ("Pantalla Inf Note 30 Pro / Camon 20", 1, "Infinix", "Note 30 Pro", "", '["Note 30 Pro", "Camon 20"]', 0, 0, 2, 2),
        ("Pantalla Y9 2019 Huawei", 1, "Huawei", "Y9 2019", "", '[]', 0, 0, 1, 2),
        ("Pantalla iPhone 11 Incell", 1, "Apple", "iPhone 11", "Incell", '[]', 0, 0, 2, 2),
        ("Pantalla Infinix Note 12/11", 1, "Infinix", "Note 12", "", '["Note 12", "Note 11"]', 0, 0, 2, 2),
        ("Pantalla Samsung A16 4G", 1, "Samsung", "A16 4G", "", '[]', 0, 0, 5, 2),
    ]
    count = 0
    for p in products:
        try:
            conn.execute("INSERT OR IGNORE INTO products (name, category_id, brand, model, variant, compatibility, price_cost, price_sale, stock, min_stock) VALUES (?,?,?,?,?,?,?,?,?,?)", p)
            count += 1
        except:
            pass
    print(f"  Products: {count} imported")
    return count

def main():
    conn = init_db()
    print("Importing data...")
    existing = conn.execute("SELECT COUNT(*) FROM services").fetchone()[0]
    if existing > 0:
        print(f"  DB already has {existing} services - clearing for fresh import")
        for t in ['inventory_movements', 'sales', 'services', 'products']:
            conn.execute(f"DELETE FROM {t}")
        conn.commit()

    import_products(conn)
    import_from_excel(conn, os.path.join(os.path.dirname(__file__), 'REGISTRO OFICIAL.xlsx'), "REGISTRO OFICIAL")
    import_from_excel(conn, r'C:\Users\ROBER\Downloads\Sistema_Reparacion_Atencion_Cliente (1).xlsx', "Sistema_Reparacion")
    import_from_excel(conn, r'C:\Users\ROBER\Downloads\Sistema_Reparacion_Atencion_Cliente.xlsx', "Sistema_Reparacion (alt)")

    conn.commit()
    total_services = conn.execute("SELECT COUNT(*) FROM services").fetchone()[0]
    total_products = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    print(f"\nTotal: {total_services} services, {total_products} products")
    conn.close()

    export_to_json()
    print("Done! Database ready for Tauri app.")

if __name__ == '__main__':
    main()
